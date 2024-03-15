from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

# Replace with your desired file path
output_file = "states_and_hrefs.xlsx"

driver = webdriver.Chrome()
driver.get("https://www.angi.com/companylist/lawn-and-yard-work.htm")

# Find all states
states = driver.find_elements(By.XPATH, "/html/body/div[2]/main/div[2]/div/div/ul[3]/li/a")

# Create a new workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Href"

# Write extracted data to Excel
row_count = 2
# Extract state names and hrefs
for state in states:
    state_name = state.text
    href = state.get_attribute("href")
    print(f"State: {state_name}, Href: {href}")
    sheet.cell(row=row_count, column=1).value = state_name
    sheet.cell(row=row_count, column=2).value = href
    row_count += 1


# Save the workbook
workbook.save(output_file)

driver.quit()
print(f"States and hrefs saved successfully to {output_file}")
